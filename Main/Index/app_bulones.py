# Variables
#Defino forma de importación del excel
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook

continuar = 0



#Defino ubicación del archivo

# los diametros posibles de seleccionar son = 1, 2, 3, 4, 5
# los largos posibles de seleccionar son = 10, 20, 30, 40, 50
# las terminaciones posibles son = "negro", "zincado"

# El usuario debe ingresar los atributos del bulon
# para verificar si el mismo esta disponible

#Ingresa los datos y verifica que los mismos son correctos"

def verificacion_parametros():
    global diametro_input
    global largo_input
    global terminacion_input
    while True: 
        while True:
            try:
                diametro_input = int(input("Introduzca el diametro requerido: "))
            except ValueError:
                print("Debes escribir un numero")
                continue
            if diametro_input > 5 or diametro_input == 0:
                print("Ponga un numero válido 1, 2, 3, 4, 5")
            else:
                print("Pusiste numero válido")
                break
            continue

        while True:
            try:
                largo_input = int(input("Introduzca el largo requerido: "))
            except ValueError:
                print("Debes escribir un numero")
                continue
            if largo_input == 10 and largo_input == 20 and largo_input == 30:
                print("Ponga un numero válido 10, 20, 30, 40, 50")
            else:
                print("Pusiste numero válido")
                break
            continue

        while True:
            terminacion_input = (input("Introduzca la terminación requerida: "))
            if terminacion_input != "negro" and terminacion_input != "zincado":
                print("Ponga una terminación valida negro o zincado")
            else:
                print("Pusiste terminación válida")
                break
            continue
        
        print("Usted selecciono un bulón:")
        print(("Diametro:  ") + str(diametro_input))
        print(("Largo:  ") + str(largo_input))
        print(("Terminación:") + terminacion_input)
        print("Si el bulón que selecciono es correcto presione 1")
        print("De no ser correcto presione 0")
        paso_1 = int(input("Desea continuar "))
        if paso_1 == 0:
            print("Vuelva a seleccionar el bulón")        
        else:
            break
        continue

def cantidad_requerida():
    while True:
        global cantidad_requerida
        try:
            cantidad_requerida = int(input("Introduzca la cantidad requerida: "))
        except ValueError:
            print("Debes escribir un numero")
        else:
            break
        continue


def genero_superdescriptor():
        global superd
        if terminacion_input == "negro":
            superd_terminacion = 1 
        if terminacion_input == "zincado":
            superd_terminacion = 2
        superd = str(diametro_input)+str(largo_input)+str(superd_terminacion)


def agregar_al_carrito():
    filesheet = ("C:\Desarrollo\GIT\B-BUL\B-BUL\Main\Index\data_base.xlsx")
    wb = load_workbook(filesheet)
    hoja_carrito = wb.create_sheet("carrito")
    hoja_carrito["A1"] = "superdescriptor"
    hoja_carrito["A2"] = superd
    hoja_carrito["B1"] = "cantidad"
    hoja_carrito["B2"] = cantidad_requerida
    wb.save(filesheet)


def verificar_stock():
    global fila
    global stock
    global cantidad_requerida
    #cargamos el archivo
    filesheet = ("C:\Desarrollo\GIT\B-BUL\B-BUL\Main\Index\data_base.xlsx")
    wb = load_workbook(filesheet)
    sheet = wb['stock'] #cargamos la hoja

    # Busco con el superdescriptor
    for cell in sheet["E"]:
        if cell.value == int(superd):
            stock = int(sheet[f"D{cell.row}"].value)
            fila = cell.row
            if stock >= cantidad_requerida:
                print("Tenemos esa cantidad")
            else:
                print("El stock disponible es", stock)
                cantidad_requerida = int(input("Seleccione nueva cantidad:"))
                break
            continue

def ajustar_stock():
    filesheet = ("C:\Desarrollo\GIT\B-BUL\B-BUL\Main\Index\data_base.xlsx")
    wb = load_workbook(filesheet)
    # Creo la etiqueta de la celda que voy a modificar: 
    celda = str("D"+str(fila))
#    print(celda)
# Seleccionamos el archivo
    sheet = wb.active
    sheet[celda] = stock-cantidad_requerida
    # Guardamos el archivo con los cambios
    wb.save(filesheet)

def sumar_al_carrito():
    filesheet = ("C:\Desarrollo\GIT\B-BUL\B-BUL\Main\Index\data_base.xlsx")
    wb = load_workbook(filesheet)
    item_agregado = wb["carrito"]
    item_agregado["A3"] = 123
    item_agregado["B3"] = 123
    wb.save(filesheet)

#def celda_correlativa():
    cuenta_celdaA = 1
    cuenta_celdaB = 1
    while True:
        if continuar == 1:
            cuenta_celdaA += 1
            cuenta_celdaB += 1
            break
            continue
        mem_n1 = str("A"+str(cuenta_celdaA))
        mem_n2 = str("B"+str(cuenta_celdaB))
        print(mem_n1)
        print(mem_n2)
        break
        continue


##############################################################################################################################
#                                                     ATENCION                                                                           #
#                                                                                                                            #
#                                       Empieza la ejecución del programa                                                    #
#                                                                                                                            #
#                                                                                                                            #
##############################################################################################################################

# Mensaje de bienvenida
print("Buen día! Gracias por elegirnos")
print("Seleccione el bulon que desea comprar")

# Permite al usuario instertar parámetros y verifica que sean correctos

verificacion_parametros()
cantidad_requerida()
genero_superdescriptor()
verificar_stock()
agregar_al_carrito()
while True:
    continuar = (input("Desea comprar otro producto?: "))
    if continuar == "si":
        verificacion_parametros()
        cantidad_requerida()
        genero_superdescriptor()
        verificar_stock()
        sumar_al_carrito()
    else:
        ajustar_stock()
        print("Usted compró un bulón:")
        print(("Diametro:  ") + str(diametro_input))
        print(("Largo:  ") + str(largo_input))
        print(("Terminación:") + terminacion_input)
        print ("Vuelva pronto")
        break
    continue
            






